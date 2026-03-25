from flask import Flask, request, send_file, render_template_string
import pandas as pd
import numpy as np
import openpyxl
import io
import os

app = Flask(__name__)

HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8" />
<meta name="viewport" content="width=device-width, initial-scale=1.0"/>
<title>Reconciliation Processor</title>
<link href="https://fonts.googleapis.com/css2?family=Syne:wght@400;600;800&family=DM+Sans:wght@300;400;500&display=swap" rel="stylesheet"/>
<style>
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

  :root {
    --bg: #e6f0fa;        /* sky blue background */
    --surface: #ffffff;    /* white card */
    --border: #bdd4e7;
    --accent: #3b82f6;     /* sky blue accent */
    --accent2: #38bdf8;    /* lighter sky blue */
    --text: #1e293b;
    --muted: #5a6e85;
    --success: #4ade80;
  }

  body {
    background: var(--bg);
    color: var(--text);
    font-family: 'DM Sans', sans-serif;
    min-height: 100vh;
    display: flex;
    align-items: center;
    justify-content: center;
    padding: 2rem;
    overflow-x: hidden;
  }

  body::before {
    content: '';
    position: fixed;
    top: -40%;
    left: -20%;
    width: 70vw;
    height: 70vw;
    background: radial-gradient(circle, rgba(59,130,246,0.07) 0%, transparent 70%);
    pointer-events: none;
    z-index: 0;
  }

  body::after {
    content: '';
    position: fixed;
    bottom: -30%;
    right: -10%;
    width: 50vw;
    height: 50vw;
    background: radial-gradient(circle, rgba(56,189,248,0.05) 0%, transparent 70%);
    pointer-events: none;
    z-index: 0;
  }

  .card {
    position: relative;
    z-index: 1;
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: 20px;
    padding: 3rem 3.5rem;
    width: 100%;
    max-width: 520px;
    box-shadow: 0 0 60px rgba(59,130,246,0.1);
  }

  .badge {
    display: inline-block;
    font-family: 'Syne', sans-serif;
    font-size: 0.65rem;
    font-weight: 600;
    letter-spacing: 0.15em;
    text-transform: uppercase;
    color: var(--accent);
    border: 1px solid rgba(59,130,246,0.3);
    background: rgba(59,130,246,0.08);
    padding: 0.3rem 0.8rem;
    border-radius: 100px;
    margin-bottom: 1.2rem;
  }

  h1 {
    font-family: 'Syne', sans-serif;
    font-size: 2rem;
    font-weight: 800;
    line-height: 1.15;
    margin-bottom: 0.5rem;
    background: linear-gradient(135deg, var(--text) 0%, var(--muted) 100%);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
  }

  .subtitle {
    font-size: 0.88rem;
    color: var(--muted);
    margin-bottom: 2.5rem;
    line-height: 1.6;
  }

  .field {
    margin-bottom: 1.5rem;
  }

  label {
    display: flex;
    align-items: center;
    gap: 0.5rem;
    font-size: 0.8rem;
    font-weight: 500;
    color: var(--muted);
    text-transform: uppercase;
    letter-spacing: 0.08em;
    margin-bottom: 0.6rem;
  }

  label .dot {
    width: 6px;
    height: 6px;
    border-radius: 50%;
    background: var(--accent);
    display: inline-block;
  }

  label .dot.red { background: var(--accent2); }

  .file-box {
    position: relative;
    border: 1.5px dashed var(--border);
    border-radius: 12px;
    padding: 1.2rem 1.4rem;
    cursor: pointer;
    transition: border-color 0.2s, background 0.2s;
    background: rgba(59,130,246,0.015);
  }

  .file-box:hover {
    border-color: var(--accent);
    background: rgba(59,130,246,0.04);
  }

  .file-box input[type="file"] {
    position: absolute;
    inset: 0;
    opacity: 0;
    cursor: pointer;
    width: 100%;
    height: 100%;
  }

  .file-placeholder {
    display: flex;
    align-items: center;
    gap: 0.8rem;
    font-size: 0.88rem;
    color: var(--muted);
    pointer-events: none;
  }

  .file-icon {
    font-size: 1.3rem;
    flex-shrink: 0;
  }

  .file-name {
    font-size: 0.85rem;
    color: var(--text);
    margin-top: 0.3rem;
    font-style: italic;
    display: none;
  }

  .divider {
    height: 1px;
    background: var(--border);
    margin: 2rem 0;
  }

  .btn {
    width: 100%;
    padding: 1rem;
    border: none;
    border-radius: 12px;
    font-family: 'Syne', sans-serif;
    font-size: 0.95rem;
    font-weight: 600;
    letter-spacing: 0.04em;
    cursor: pointer;
    background: linear-gradient(135deg, var(--accent), var(--accent2));
    color: #fff;
    transition: opacity 0.2s, transform 0.15s;
    position: relative;
    overflow: hidden;
  }

  .btn:hover { opacity: 0.9; transform: translateY(-1px); }
  .btn:active { transform: translateY(0); }
  .btn:disabled { opacity: 0.4; cursor: not-allowed; transform: none; }

  /* Error status styling - hidden initially */
  .status {
    margin-top: 1.2rem;
    padding: 0.9rem 1.2rem;
    border-radius: 10px;
    font-size: 0.85rem;
    display: none;
    background: rgba(56,189,248,0.08);
    border: 1px solid rgba(56,189,248,0.2);
    color: var(--accent2);
  }

  /* When we want to show it, we set display: block via JS */
  .file-box.filled { border-color: rgba(74,222,128,0.4); background: rgba(74,222,128,0.03); }
</style>
</head>
<body>
<div class="card">
  <div class="badge">Reconciliation Tool</div>
  <h1>Process &amp;<br/>Export Data</h1>
  <p class="subtitle">Upload your main data file and lookup file to generate the formatted Excel output.</p>

  <form id="uploadForm">
    <div class="field">
      <label><span class="dot"></span> Main File</label>
      <div class="file-box" id="box1">
        <input type="file" name="main_file" accept=".xlsx,.xls,.html,.htm" id="file1" required />
        <div class="file-placeholder">
          <span class="file-icon">📄</span>
          <div>
            <div>HTML / XLS / XLSX</div>
            <div class="file-name" id="name1"></div>
          </div>
        </div>
      </div>
    </div>

    <div class="field">
      <label><span class="dot red"></span> Lookup File</label>
      <div class="file-box" id="box2">
        <input type="file" name="lookup_file" accept=".xlsx,.xls" id="file2" required />
        <div class="file-placeholder">
          <span class="file-icon">🔍</span>
          <div>
            <div>XLSX only</div>
            <div class="file-name" id="name2"></div>
          </div>
        </div>
      </div>
    </div>

    <div class="divider"></div>

    <button type="submit" class="btn" id="submitBtn">Generate Excel ↗</button>

    <div class="status" id="errorStatus"></div>
  </form>
</div>

<script>
  function bindFile(inputId, nameId, boxId) {
    document.getElementById(inputId).addEventListener('change', function() {
      const box = document.getElementById(boxId);
      const nameEl = document.getElementById(nameId);
      if (this.files.length) {
        nameEl.textContent = this.files[0].name;
        nameEl.style.display = 'block';
        box.classList.add('filled');
      }
    });
  }
  bindFile('file1', 'name1', 'box1');
  bindFile('file2', 'name2', 'box2');

  document.getElementById('uploadForm').addEventListener('submit', async function(e) {
    e.preventDefault();
    const btn = document.getElementById('submitBtn');
    const errorEl = document.getElementById('errorStatus');

    btn.disabled = true;
    errorEl.style.display = 'none';
    errorEl.textContent = ''; // clear previous message

    const formData = new FormData(this);

    try {
      const res = await fetch('/process', { method: 'POST', body: formData });
      if (!res.ok) {
        const err = await res.json();
        throw new Error(err.error || 'Server error');
      }
      const blob = await res.blob();
      const url = URL.createObjectURL(blob);
      const a = document.createElement('a');
      a.href = url;
      a.download = 'Final_Output.xlsx';
      a.click();
      URL.revokeObjectURL(url);
    } catch (err) {
      errorEl.textContent = '⚠ ' + err.message;
      errorEl.style.display = 'block';
    } finally {
      btn.disabled = false;
    }
  });
</script>
</body>
</html>
"""

@app.route("/")
def index():
    return render_template_string(HTML)


@app.route("/process", methods=["POST"])
def process():
    try:
        main_file = request.files.get("main_file")
        lookup_file = request.files.get("lookup_file")

        if not main_file or not lookup_file:
            return {"error": "Both files are required."}, 400

        main_bytes = main_file.read()
        lookup_bytes = lookup_file.read()

        # =============================
        # 2️⃣ MAIN FILE READ (HTML या Excel)
        # =============================
        try:
            tables = pd.read_html(io.BytesIO(main_bytes))
            df = tables[0]
        except Exception:
            df = pd.read_excel(io.BytesIO(main_bytes), dtype={"UTR NO": str})

        # =============================
        # 3️⃣ HEADER FIX
        # =============================
        df.columns = df.iloc[0]
        df = df[1:].reset_index(drop=True)

        # =============================
        # 4️⃣ CLEAN COLUMN NAMES
        # =============================
        df.columns = df.columns.str.strip()

        # =============================
        # 5️⃣ NUMERIC FIX
        # =============================
        numeric_cols = [
            "NET BILL AMT.", "SPONSER_AMOUNT", "CLAIM AMOUNT",
            "RECEIVED AMOUNT", "TDS AMOUNT", "WRITEOFF AMOUNT",
            "PATIENT AMOUNT", "PROCESSING FEE", "LEGITIMATE DISCOUNT",
            "DISALLOWANCE AMOUNT", "Total"
        ]
        for col in numeric_cols:
            if col in df.columns:
                df[col] = df[col].astype(str).str.replace(",", "")
                df[col] = pd.to_numeric(df[col], errors="coerce")

        # =============================
        # 6️⃣ DATE FIX
        # =============================
        date_columns = ["FILE_SUBMISSION_DT", "UTR DATE", "INVOICE DATE", "RECONCILED DATE"]
        for col in date_columns:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors="coerce", dayfirst=True)

        # =============================
        # 7️⃣ FILTER UNIT_NAME
        # =============================
        if "UNIT_NAME" in df.columns:
            df = df[~df["UNIT_NAME"].isin(["Zynova", "---END---"])]

        # =============================
        # 8️⃣ UTR NO FINAL FIX
        # =============================
        if "UTR NO" in df.columns:
            df["UTR NO"] = df["UTR NO"].fillna("").astype(str)
            df["UTR NO"] = df["UTR NO"].replace("nan", "").str.strip()

        # =============================
        # 9️⃣ VISIT COLUMN
        # =============================
        if "VISIT_ID" in df.columns:
            df["VISIT"] = df["VISIT_ID"].astype(str).str[:2]
            df["VISIT"] = df["VISIT"].replace("ER", "OP")
            df["VISIT_ID"] = df["VISIT_ID"].astype(str).str.replace("^ER", "OP", regex=True)

        # =============================
        # 🔟 LOOKUP FILE
        # =============================
        lookup_df = pd.read_excel(io.BytesIO(lookup_bytes))
        lookup_df.columns = lookup_df.columns.str.strip()
        lookup_df = lookup_df.iloc[:, 1:4]
        lookup_df.columns = ["SPONSOR", "Existing", "Payer"]
        payer_map    = lookup_df.drop_duplicates(subset="SPONSOR").set_index("SPONSOR")["Payer"]
        existing_map = lookup_df.drop_duplicates(subset="SPONSOR").set_index("SPONSOR")["Existing"]
        if "SPONSOR" in df.columns:
            df["Payer"]    = df["SPONSOR"].map(payer_map).fillna("NA")
            df["Existing"] = df["SPONSOR"].map(existing_map).fillna("NA")

        # =============================
        # 1️⃣1️⃣ TOTAL COLUMN
        # =============================
        amount_cols = [
            "RECEIVED AMOUNT", "TDS AMOUNT", "WRITEOFF AMOUNT",
            "PATIENT AMOUNT", "PROCESSING FEE", "LEGITIMATE DISCOUNT",
            "DISALLOWANCE AMOUNT"
        ]
        for col in amount_cols:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        df["Total"] = df[[c for c in amount_cols if c in df.columns]].sum(axis=1)

        # =============================
        # 1️⃣2️⃣ FINAL COLUMN ORDER
        # =============================
        final_cols = [
            "UNIT_NAME", "RECONCILED DATE", "VISIT", "VISIT_ID", "ADMISSION NUMBER",
            "MRNO", "PATIENT NAME", "INVOICE_NO", "INVOICE DATE", "Payer", "Existing",
            "SPONSOR", "UTR NO", "UTR DATE", "NET BILL AMT.", "SPONSER_AMOUNT",
            "CLAIM AMOUNT", "RECEIVED AMOUNT", "TDS AMOUNT", "WRITEOFF AMOUNT",
            "PATIENT AMOUNT", "PROCESSING FEE", "LEGITIMATE DISCOUNT",
            "DISALLOWANCE AMOUNT", "Total", "REMARKS", "FILE_SUBMISSION_DT",
            "IS RESUBMISION", "ADMITTING DR.", "SPECIALITY"
        ]
        final_cols = [c for c in final_cols if c in df.columns]
        final_df = df[final_cols]

        # =============================
        # 1️⃣3️⃣ SAVE → OPENPYXL FORMATTING
        # =============================
        temp_buf = io.BytesIO()
        final_df.to_excel(temp_buf, index=False)
        temp_buf.seek(0)

        wb = openpyxl.load_workbook(temp_buf)
        ws = wb.active

        # DATE FORMAT APPLY
        for col_name in date_columns:
            if col_name in final_df.columns:
                col_index = list(final_df.columns).index(col_name) + 1
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_index)
                    if cell.value:
                        cell.number_format = "DD-MM-YYYY"

        # NUMBER FORMAT FIX
        for col in range(1, ws.max_column + 1):
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0"

        # =============================
        # 1️⃣4️⃣ RETURN FILE
        # =============================
        out_buf = io.BytesIO()
        wb.save(out_buf)
        out_buf.seek(0)

        return send_file(
            out_buf,
            as_attachment=True,
            download_name="Final_Output.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return {"error": str(e)}, 500


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
